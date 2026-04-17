# -*- coding: utf-8 -*-

"""Generate golgg/source/teams-All.xlsx from gol.gg team list tournament pages.

This automates the manual "Copy table to clipboard" workflow and writes the
same style sheet expected by step002_info_teams.py:
- worksheet name: teams-All
- column A: Team name (with hyperlink to team stats page)
- column B: Season label (e.g. S16)
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup

from golgg.pipeline.common import log_step_end, log_step_start
from golgg.pipeline.ingestao.rendered_page import fetch_rendered_html
from golgg.pipeline.ingestao.teams_all_xlsx import extract_headers, extract_rows, find_teams_table


BASE_URL = "https://gol.gg"
DEFAULT_TOURNAMENT_URLS = [
    "https://gol.gg/teams/list/season-ALL/split-ALL/tournament-CBLOL%202026%20Split%201/",
    "https://gol.gg/teams/list/season-ALL/split-ALL/tournament-CBLOL%20Cup%202026/",
]
DEFAULT_OUTPUT = Path("golgg/data/bronze/teams-All.xlsx")
DEFAULT_SHEET = "teams-All"


def write_workbook(headers: list[str], rows: list[dict[str, object]], output_path: Path, sheet_name: str) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    effective_headers = list(headers)
    if "Tournament" not in effective_headers:
        effective_headers.append("Tournament")

    for col, header in enumerate(effective_headers, start=1):
        worksheet.cell(row=1, column=col, value=header)

    for idx, row_data in enumerate(rows, start=2):
        values = row_data["values"]
        for col, value in enumerate(values, start=1):
            worksheet.cell(row=idx, column=col, value=value)

        tournament_col = effective_headers.index("Tournament") + 1
        worksheet.cell(row=idx, column=tournament_col, value=row_data.get("tournament", "N/A"))

        team_url = str(row_data.get("team_url", "")).strip()
        if team_url:
            worksheet.cell(row=idx, column=1).hyperlink = team_url

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate teams-All.xlsx from gol.gg teams list pages")
    parser.add_argument(
        "--tournament-url",
        action="append",
        dest="tournament_urls",
        default=[],
        help="Tournament teams list URL. Can be repeated. Defaults to CBLOL 2026 Split 1 and CBLOL Cup 2026.",
    )
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output XLSX path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    start_time = log_step_start("step001_generate_teams_all_xlsx")

    tournament_urls = args.tournament_urls or DEFAULT_TOURNAMENT_URLS
    all_rows: list[dict[str, object]] = []
    headers: list[str] = []

    for url in tournament_urls:
        print(f"[step01] loading {url}", flush=True)
        html = fetch_rendered_html(url)
        soup = BeautifulSoup(html, "html.parser")
        table = find_teams_table(soup)
        if table is None:
            raise RuntimeError(f"Could not find teams table in {url}")

        current_headers = extract_headers(table)
        if not headers:
            headers = current_headers
        elif current_headers != headers:
            raise RuntimeError(
                "Teams table headers changed across tournaments; aborting to avoid invalid workbook format"
            )

        rows = extract_rows(table, url)
        all_rows.extend(rows)
        print(f"[step01] extracted rows={len(rows)}", flush=True)

    if not headers:
        raise RuntimeError("No headers found while generating teams-All.xlsx")

    output_path = Path(args.output)
    write_workbook(headers, all_rows, output_path, args.sheet)
    print(f"[step01] wrote {output_path} with rows={len(all_rows)}", flush=True)

    log_step_end("step001_generate_teams_all_xlsx", start_time)


if __name__ == "__main__":
    main()

