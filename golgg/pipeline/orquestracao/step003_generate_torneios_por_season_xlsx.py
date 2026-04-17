# -*- coding: utf-8 -*-

"""Generate golgg/source/torneios_por_season.xlsx from gol.gg tournament list."""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

from golgg.pipeline.common import log_step_end, log_step_start
from golgg.pipeline.ingestao.rendered_page import fetch_rendered_html
from golgg.pipeline.ingestao.torneios_por_season import normalize_text, read_tournament_rows


TOURNAMENT_LIST_URL = "https://gol.gg/tournament/list/"
DEFAULT_OUTPUT = Path("golgg/data/bronze/torneios_por_season.xlsx")
DEFAULT_SHEET = "torneios"
DEFAULT_TARGETS = [
    "CBLOL 2026 Split 1",
    "CBLOL Cup 2026",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate torneios_por_season.xlsx from gol.gg")
    parser.add_argument(
        "--target",
        action="append",
        dest="targets",
        default=[],
        help="Tournament name to include. Can be repeated. Defaults to CBLOL 2026 Split 1 and CBLOL Cup 2026.",
    )
    parser.add_argument("--season", default="16", help="Season value for the SEASON column (default: 16)")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output XLSX path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name")
    return parser.parse_args()


def write_workbook(rows: list[dict[str, str]], season: str, output_path: Path, sheet_name: str) -> None:
    headers = [
        "NAME",
        "REGION",
        "NUMBER OF GAMES",
        "GAME DURATION",
        "FIRST GAME",
        "LAST GAME",
        "SEASON",
    ]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col, value=header)

    for row_idx, row in enumerate(rows, start=2):
        worksheet.cell(row=row_idx, column=1, value=row["NAME"])
        worksheet.cell(row=row_idx, column=2, value=row["REGION"])
        worksheet.cell(row=row_idx, column=3, value=row["NUMBER OF GAMES"])
        worksheet.cell(row=row_idx, column=4, value=row["GAME DURATION"])
        worksheet.cell(row=row_idx, column=5, value=row["FIRST GAME"])
        worksheet.cell(row=row_idx, column=6, value=row["LAST GAME"])
        worksheet.cell(row=row_idx, column=7, value=season)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    start_time = log_step_start("step003_generate_torneios_por_season_xlsx")

    targets = args.targets or DEFAULT_TARGETS
    target_set = {normalize_text(target) for target in targets}

    html = fetch_rendered_html(TOURNAMENT_LIST_URL)
    all_rows = read_tournament_rows(html)
    filtered_rows = []
    for row in all_rows:
        normalized_name = normalize_text(row["NAME"])
        if any(target in normalized_name for target in target_set):
            filtered_rows.append(row)

    if not filtered_rows:
        raise RuntimeError(f"No tournaments matched targets: {targets}")

    output_path = Path(args.output)
    write_workbook(filtered_rows, str(args.season), output_path, args.sheet)

    print(f"[step02] wrote {output_path} with rows={len(filtered_rows)}", flush=True)
    log_step_end("step003_generate_torneios_por_season_xlsx", start_time)


if __name__ == "__main__":
    main()

